import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("Error: openpyxl module not found.")
    sys.exit(1)

def get_color(cell):
    if cell.fill and cell.fill.patternType == 'solid':
        return cell.fill.start_color.index # or rgb
    return None

def main():
    file_path = "Stoffice_Berechnungstabelle_Datenpool_sst_260104.xlsx"
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False) # data_only=False to get formulas
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    print("--- START ANALYSIS ---")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name}")
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                    
                # Extract Color Info safely
                color_info = "None"
                is_yellow = False
                is_orange = False
                
                if cell.fill and cell.fill.patternType == 'solid':
                    fg = cell.fill.start_color
                    # Attempt to get hex
                    if hasattr(fg, 'rgb') and fg.rgb:
                        color_info = str(fg.rgb)
                        # Common Yellows: FFFF00, FFFFFF00. Common Oranges: FFC000, FFFFC000
                        if 'FF00' in color_info: is_yellow = True
                        if 'C000' in color_info or 'FFC000' in color_info: is_orange = True
                    elif hasattr(fg, 'index'):
                        color_info = f"Index:{fg.index}"
                        if fg.index == 5: is_yellow = True # Standard Yellow Index
                
                # Get Label (Left neighbor)
                label = "Unknown"
                if cell.col_idx > 1:
                    neighbor = ws.cell(row=cell.row, column=cell.col_idx-1)
                    if neighbor.value:
                        label = str(neighbor.value)
                    # Try 2 columns left if immediate is empty (common in formatting)
                    elif cell.col_idx > 2:
                        neighbor_2 = ws.cell(row=cell.row, column=cell.col_idx-2)
                        if neighbor_2.value:
                            label = str(neighbor_2.value)

                val_str = str(cell.value)
                
                # Check explicit user instruction mapping (Yellow=Input, Orange=Formula)
                category = "OTHER"
                if is_yellow: category = "INPUT (Yellow)"
                elif is_orange: category = "FORMULA (Orange)"
                elif "=" in val_str: category = "FORMULA (Implicit)"

                if category != "OTHER":
                    print(f"[{category}] Cell: {cell.coordinate} | Label: {label} | Val: {val_str} | Color: {color_info}")


    print("--- END ANALYSIS ---")

if __name__ == "__main__":
    main()
